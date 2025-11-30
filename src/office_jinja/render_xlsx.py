import re
import os
from pydantic import BaseModel
import logging
from typing import Generator
from openpyxl.worksheet.worksheet import Worksheet
from .models import Context, Template, TemplateFile
from openpyxl import load_workbook
from xltpl.writerx import BookWriter
from PIL import Image

VAR_PATTERN = re.compile(r"{{\s*([\w\.]+)\s*}}")
FOR_PATTERN = re.compile(r"{%-?\s*for\s*([\w\.]+)\s*in\s*([\w\.]+)\s*%}")
END_FOR_PATTERN = re.compile(r"{%-?\s*endfor\s*%}")


class VarPosition(BaseModel):
    row: int
    column: int
    value: str
    type: str


def resolve_var(name: str, context: dict) -> str:
    """支持 a.b.c 形式的取值"""
    parts = name.split(".")
    val = context
    for p in parts:
        if p not in val:
            return ""
        val = val[p]
    return str(val)


def render_all_xlsx_from_template(template: Template) -> None:
    """
    渲染所有xlsx模板
    """
    for template_file in template.templates:
        if not template_file.path.endswith(".xlsx"):
            continue
        render_xlsx(template_file, template.context)


def iter_all_pattern_from_sheet(sheet: Worksheet) -> Generator[VarPosition, None, None]:
    """
    遍历工作表中的所有占位符
    """
    pattern_map = {
        "for": FOR_PATTERN,
        "var": VAR_PATTERN,
        "endfor": END_FOR_PATTERN,
    }
    for row in sheet.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            for pattern_type, pattern in pattern_map.items():
                match = pattern.search(cell.value)
                if match:
                    yield VarPosition(
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                        type=pattern_type
                    )


def read_images(context: Context) -> dict:
    """
    读取图片
    """
    images = {}
    for pic_file, pic_path in context.pic.items():
        if not os.path.exists(pic_path):
            logging.error(f"Picture file {pic_path} not found")
            raise FileNotFoundError(f"Picture file {pic_path} not found")
        images[pic_file] = Image.open(pic_path)
    return images


def load_data(template_file: TemplateFile, context: Context) -> dict:
    """
    加载数据
    """
    wb = load_workbook(template_file.path)
    data = []
    images = read_images(context)
    for sheet in wb.sheetnames:
        data.append({
            "sheet_name": sheet,
            "tpl_name": sheet,
            **context.data,
            **images
        })
    return data


def render_xlsx(template_file: TemplateFile, context: Context) -> None:
    """
    渲染单个xlsx模板
    """
    bw = BookWriter(template_file.path)
    data = load_data(template_file, context)
    bw.render_book(data)
    bw.save(template_file.output_path)
